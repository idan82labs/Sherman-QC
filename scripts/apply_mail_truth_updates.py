#!/opt/homebrew/bin/python3.11
"""
Apply mail-delivered truth updates to the bend corpus without rebuilding it.

This script:
- parses the shared XLSX truth table
- records a normalized JSON annotation file
- updates per-part metadata/labels
- links newly supplied drawings into the corpus
- refreshes inventory and confirmation views

It intentionally preserves existing evaluation outputs under data/bend_corpus/evaluation.
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


ROOT = Path("/Users/idant/82Labs/Sherman QC/Full Project")
CORPUS_ROOT = ROOT / "data" / "bend_corpus"
PARTS_DIR = CORPUS_ROOT / "parts"

MAIL_XLSX = Path("/Users/idant/Downloads/נתונים עבור חלקים.xlsx")
MAIL_DRAWINGS = [
    Path("/Users/idant/Downloads/שרטוט  37361005 1.PDF"),
    Path("/Users/idant/Downloads/שרטוט  11979003.PDF"),
    Path("/Users/idant/Downloads/שרטוט 44211000.pdf"),
    Path("/Users/idant/Downloads/שרטוט 47266000 1.pdf"),
    Path("/Users/idant/Downloads/שרטוט 48963008.pdf"),
    Path("/Users/idant/Downloads/שרטוט 47959001.pdf"),
    Path("/Users/idant/Downloads/שרטוט 49001000.pdf"),
    Path("/Users/idant/Downloads/שרטוט 49024000.PDF"),
    Path("/Users/idant/Downloads/שרטוט 47479000.PDF"),
]


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _dump_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _part_key_from_text(text: str) -> Optional[str]:
    matches = re.findall(r"\d{6,}", text or "")
    if not matches:
        return None
    return max(matches, key=len)


def _file_record(path: Path) -> Dict[str, Any]:
    st = path.stat()
    return {
        "name": path.name,
        "path": str(path),
        "mtime": datetime.fromtimestamp(st.st_mtime).isoformat(),
        "size_bytes": st.st_size,
    }


def _safe_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned or "file"


def _unlink_or_remove(path: Path) -> None:
    if not path.exists() and not path.is_symlink():
        return
    if path.is_symlink() or path.is_file():
        path.unlink()
    else:
        raise RuntimeError(f"Unexpected directory collision at {path}")


def _ensure_symlink(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    _unlink_or_remove(dst)
    os.symlink(src, dst)


def _parse_mail_truth() -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(MAIL_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: Dict[str, Dict[str, Any]] = {}
    for row in range(7, ws.max_row + 1):
        part_id = ws.cell(row, 4).value
        if part_id in (None, ""):
            continue
        part_key = _part_key_from_text(str(part_id))
        if not part_key:
            continue
        expected_raw = ws.cell(row, 5).value
        expected_bends: Optional[int] = None
        expected_note: Optional[str] = None
        if isinstance(expected_raw, (int, float)):
            expected_bends = int(expected_raw)
        elif isinstance(expected_raw, str) and expected_raw.strip():
            expected_note = expected_raw.strip()
        rows[part_key] = {
            "part_key": part_key,
            "mail_source_xlsx": str(MAIL_XLSX),
            "expected_total_bends": expected_bends,
            "expected_bends_note": expected_note,
            "scan_type_note": str(ws.cell(row, 6).value or "").strip() or None,
            "completed_bends_note": str(ws.cell(row, 7).value or "").strip() or None,
            "drawing_available_note": str(ws.cell(row, 8).value or "").strip() or None,
            "quality_or_comments": str(ws.cell(row, 9).value or "").strip() or None,
        }
    return rows


def _mail_drawing_map() -> Dict[str, List[Path]]:
    out: Dict[str, List[Path]] = {}
    for path in MAIL_DRAWINGS:
        if not path.exists():
            continue
        part_key = _part_key_from_text(path.name)
        if not part_key:
            continue
        out.setdefault(part_key, []).append(path)
    return out


def _refresh_labels(
    labels: Dict[str, Any],
    truth: Dict[str, Any],
    metadata: Dict[str, Any],
) -> Dict[str, Any]:
    labels["expected_total_bends"] = truth.get("expected_total_bends")
    note_lines: List[str] = []
    if truth.get("scan_type_note"):
        note_lines.append(f"mail_scan_type={truth['scan_type_note']}")
    if truth.get("expected_bends_note"):
        note_lines.append(f"mail_expected_note={truth['expected_bends_note']}")
    if truth.get("quality_or_comments"):
        note_lines.append(f"mail_comment={truth['quality_or_comments']}")
    labels["notes"] = " | ".join(note_lines)

    scans_by_name = {
        str(item.get("filename")): item
        for item in (labels.get("scans") or [])
        if isinstance(item, dict) and item.get("filename")
    }
    for scan in metadata.get("scan_files") or []:
        name = scan["name"]
        row = scans_by_name.get(name)
        if row is None:
            row = {
                "filename": name,
                "state": scan.get("state") or "unknown",
                "completed_bends": None,
                "completed_bend_ids": [],
                "all_completed_in_spec": None,
                "comments": "",
            }
            labels.setdefault("scans", []).append(row)
        state = str(scan.get("state") or row.get("state") or "unknown")
        row["state"] = state
        if state == "full" and truth.get("expected_total_bends") is not None:
            row["completed_bends"] = int(truth["expected_total_bends"])
        comment_parts: List[str] = []
        if truth.get("scan_type_note"):
            comment_parts.append(f"mail_scan_type={truth['scan_type_note']}")
        if truth.get("quality_or_comments"):
            comment_parts.append(truth["quality_or_comments"])
        row["comments"] = " | ".join(part for part in comment_parts if part)
        if truth.get("quality_or_comments") and "לא תקין" in truth["quality_or_comments"] and state == "full":
            row["all_completed_in_spec"] = False
    return labels


def _link_drawings(part_key: str, drawing_files: List[Dict[str, Any]]) -> None:
    part_draw_dir = PARTS_DIR / part_key / "drawings"
    part_draw_dir.mkdir(parents=True, exist_ok=True)
    for idx, item in enumerate(drawing_files, start=1):
        src = Path(item["path"])
        dst = part_draw_dir / f"{idx:02d}_{_safe_name(src.name)}"
        _ensure_symlink(src, dst)


def _refresh_views() -> None:
    parts: List[Dict[str, Any]] = []
    for part_dir in sorted(PARTS_DIR.iterdir()):
        if not part_dir.is_dir():
            continue
        meta_path = part_dir / "metadata.json"
        if not meta_path.exists():
            continue
        meta = _load_json(meta_path)
        labels_path = part_dir / "labels.template.json"
        labels = _load_json(labels_path) if labels_path.exists() else {}

        known_expected = meta.get("expected_bends_override") is not None
        has_drawings = bool(meta.get("drawing_files"))
        has_specs = bool(meta.get("spec_files"))
        readiness = meta.get("readiness") or {}
        mail_truth = meta.get("mail_truth") or {}
        scan_type_note = str(mail_truth.get("scan_type_note") or "").strip().lower()
        expected_note = str(mail_truth.get("expected_bends_note") or "").strip()
        partial_scans = [s for s in meta.get("scan_files") or [] if s.get("state") == "partial"]
        full_scans = [s for s in meta.get("scan_files") or [] if s.get("state") == "full"]
        reasons: List[str] = []
        if not meta.get("cad_files"):
            reasons.append("missing CAD")
        if not meta.get("scan_files"):
            reasons.append("missing scans")
        if not has_drawings:
            reasons.append("no drawing linked")
        if not has_specs:
            reasons.append("no spec sheet linked")
        if not known_expected:
            reasons.append("unknown expected bend count")
        if expected_note:
            reasons.append(f"expected bends described qualitatively: {expected_note}")
        if partial_scans:
            labels_path = part_dir / "labels.template.json"
            labels = _load_json(labels_path) if labels_path.exists() else {}
            label_scans = {
                str(item.get("filename")): item
                for item in (labels.get("scans") or [])
                if isinstance(item, dict) and item.get("filename")
            }
            if any(label_scans.get(s["name"], {}).get("completed_bends") is None for s in partial_scans):
                reasons.append("partial scan completed bend count unknown")
        if scan_type_note:
            if "full" in scan_type_note and not full_scans:
                reasons.append("mail data says full but corpus has no full scan")
            if "partial" in scan_type_note and not partial_scans:
                reasons.append("mail data says partial but corpus has no partial scan")
            if "partial" in scan_type_note and "full" in scan_type_note and len(meta.get("scan_files") or []) < 2:
                reasons.append("mail data indicates both partial and full scans but corpus has fewer than two scans")
        readiness.update(
            {
                "ready_regression": bool(meta.get("cad_files") and meta.get("scan_files")),
                "ready_partial_regression": bool(partial_scans) and bool(meta.get("cad_files")),
                "ready_full_regression": bool(full_scans) and bool(meta.get("cad_files")),
                "has_drawings": has_drawings,
                "has_specs": has_specs,
                "known_expected_bends": known_expected,
                "needs_confirmation": bool(reasons),
                "reasons": reasons,
            }
        )
        meta["readiness"] = readiness
        _dump_json(meta_path, meta)

        parts.append(meta)

    inventory = {
        "generated_at": datetime.now().isoformat(),
        "source_dir": str(Path.home() / "Downloads"),
        "corpus_root": str(CORPUS_ROOT),
        "counts": {
            "parts": len(parts),
            "cad_files": sum(p["cad_count"] for p in parts),
            "scan_files": sum(p["scan_count"] for p in parts),
            "drawing_files": sum(p["drawing_count"] for p in parts),
            "spec_files": sum(p["spec_count"] for p in parts),
            "ready_regression_parts": sum(1 for p in parts if p["readiness"]["ready_regression"]),
            "ready_partial_parts": sum(1 for p in parts if p["readiness"]["ready_partial_regression"]),
            "ready_full_parts": sum(1 for p in parts if p["readiness"]["ready_full_regression"]),
            "needs_confirmation_parts": sum(1 for p in parts if p["readiness"]["needs_confirmation"]),
        },
        "parts": parts,
    }
    _dump_json(CORPUS_ROOT / "dataset_inventory.json", inventory)

    ready_regression = [
        {
            "part_key": part["part_key"],
            "expected_bends_override": part["expected_bends_override"],
            "partial_scan_count": part["partial_scan_count"],
            "full_scan_count": part["full_scan_count"],
            "drawing_count": part["drawing_count"],
            "spec_count": part["spec_count"],
            "metadata_path": str(PARTS_DIR / part["part_key"] / "metadata.json"),
            "labels_template_path": str(PARTS_DIR / part["part_key"] / "labels.template.json"),
        }
        for part in parts
        if part["readiness"]["ready_regression"]
    ]
    confirmation_queue = [
        {
            "part_key": part["part_key"],
            "expected_bends_override": part["expected_bends_override"],
            "reasons": part["readiness"]["reasons"],
            "metadata_path": str(PARTS_DIR / part["part_key"] / "metadata.json"),
        }
        for part in parts
        if part["readiness"]["needs_confirmation"]
    ]
    _dump_json(CORPUS_ROOT / "ready_regression.json", ready_regression)
    _dump_json(CORPUS_ROOT / "confirmation_queue.json", confirmation_queue)

    inventory_lines = [
        "# Bend Corpus Inventory",
        "",
        f"Generated: `{inventory['generated_at']}`",
        f"Source: `{inventory['source_dir']}`",
        f"Corpus root: `{inventory['corpus_root']}`",
        "",
        "## Summary",
        "",
        f"- Parts discovered: `{inventory['counts']['parts']}`",
        f"- CAD files linked: `{inventory['counts']['cad_files']}`",
        f"- Scan files linked: `{inventory['counts']['scan_files']}`",
        f"- Drawings linked: `{inventory['counts']['drawing_files']}`",
        f"- Specs linked: `{inventory['counts']['spec_files']}`",
        f"- Ready regression parts: `{inventory['counts']['ready_regression_parts']}`",
        f"- Ready partial-regression parts: `{inventory['counts']['ready_partial_parts']}`",
        f"- Ready full-regression parts: `{inventory['counts']['ready_full_parts']}`",
        f"- Parts needing bend-count confirmation: `{inventory['counts']['needs_confirmation_parts']}`",
        "",
        "## Parts",
        "",
        "| Part | CAD | Scans | Partial | Full | Drawings | Specs | Expected bends | Ready | Needs confirmation |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---|---|",
    ]
    for part in parts:
        inventory_lines.append(
            f"| `{part['part_key']}` | {part['cad_count']} | {part['scan_count']} | {part['partial_scan_count']} | "
            f"{part['full_scan_count']} | {part['drawing_count']} | {part['spec_count']} | "
            f"{part['expected_bends_override'] if part['expected_bends_override'] is not None else '-'} | "
            f"{'yes' if part['readiness']['ready_regression'] else 'no'} | "
            f"{'yes' if part['readiness']['needs_confirmation'] else 'no'} |"
        )
    (CORPUS_ROOT / "dataset_inventory.md").write_text("\n".join(inventory_lines) + "\n", encoding="utf-8")

    confirm_lines = ["# Confirmation Queue", ""]
    for item in confirmation_queue:
        confirm_lines.append(f"- `{item['part_key']}`: {', '.join(item['reasons'])}")
    (CORPUS_ROOT / "confirmation_queue.md").write_text("\n".join(confirm_lines) + "\n", encoding="utf-8")


def main() -> int:
    mail_truth = _parse_mail_truth()
    drawing_map = _mail_drawing_map()
    applied: List[Dict[str, Any]] = []

    truth_dir = CORPUS_ROOT / "annotations"
    truth_dir.mkdir(parents=True, exist_ok=True)
    _dump_json(truth_dir / "mail_truth_20260308.json", mail_truth)

    for part_key, truth in mail_truth.items():
        part_dir = PARTS_DIR / part_key
        meta_path = part_dir / "metadata.json"
        labels_path = part_dir / "labels.template.json"
        if not meta_path.exists() or not labels_path.exists():
            continue
        meta = _load_json(meta_path)
        labels = _load_json(labels_path)

        if truth.get("expected_total_bends") is not None:
            meta["expected_bends_override"] = int(truth["expected_total_bends"])

        new_drawings = drawing_map.get(part_key, [])
        existing_by_path = {item["path"]: item for item in meta.get("drawing_files") or []}
        drawing_files = list(meta.get("drawing_files") or [])
        for drawing in new_drawings:
            if str(drawing) in existing_by_path:
                continue
            drawing_files.append(_file_record(drawing))
        meta["drawing_files"] = drawing_files
        meta["drawing_count"] = len(drawing_files)
        meta["mail_truth"] = truth

        labels = _refresh_labels(labels, truth, meta)

        _dump_json(meta_path, meta)
        _dump_json(labels_path, labels)
        _link_drawings(part_key, drawing_files)

        applied.append(
            {
                "part_key": part_key,
                "expected_total_bends": truth.get("expected_total_bends"),
                "expected_bends_note": truth.get("expected_bends_note"),
                "scan_type_note": truth.get("scan_type_note"),
                "comment": truth.get("quality_or_comments"),
                "drawings_added": [str(p) for p in new_drawings],
            }
        )

    _dump_json(truth_dir / "mail_truth_20260308_applied.json", applied)
    _refresh_views()
    print(json.dumps({"applied": applied}, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
